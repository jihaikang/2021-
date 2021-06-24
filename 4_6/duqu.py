# import xlrd
from openpyxl import load_workbook 
import json


# cell = sheet['A:B']#记得大冒号
# 获取行列，坐标
# print(cell.row,cell.column,cell.coordinate) 
# print(cell)
# 提取数据
# for row in sheet.iter_rows(min_row=2,max_row=14614,min_col=1,max_col=2):
#     for cell in row:
#         print(cell.value) #Cell 'Sheet3'.B14529>怎么把他转化成字符串呢 加value可以打印出字来
def get_value(): 
    workbook = load_workbook(filename = 'D:\囧包子\Documents\ifma.xlsx')
    # data =('D:\囧包子\Documents\ifma.xlsx')
    print(workbook.sheetnames)
    # table1 = data.sheet_by_name
    sheet = workbook['Sheet3'] # 列表中括号
    print(sheet.dimensions)
    rows = sheet.max_row
    cols = sheet.max_column
    print(rows,cols)
    
    d=[]
    b=[]
    for j in range(3,rows+1):
        q='%s'% sheet.cell(j,1).value
        p =sheet.cell(j,2).value # 已经是这个格式就不需要再格式化了,不是而是字典的值可以是数字int
        d.append(q)
        b.append(p)
    # print(sheet.cell(3,1).value)
    # print(d)# 打印出一列的数据
    # print(b)
    zidian = {item.upper():d for item,d in zip(d,b)} # 字典生成式
    str = json.dumps(zidian)
    print(type(str))

    
    # print(zidian)
    fp = open('账号密码.txt','a',encoding='utf-8') 
    fp.write(str)
    # fp.
#     p= []
#     for i in range(3,rows+1): # SyntaxError: invalid character in identifier标识符中的无效字符 你咋永远盯上了这个for
#         d=[]
#         for j in range(1,3):
#             q='%d' % sheet.cell(i,j).value
#             d.append(q) # 这句话改一下，他的目的是干嘛
#     print(d)
# #         ap = []
# #         for k,v in d.items():
# #             if isinstance(v,float): #excel中的值默认是float,需要进行判断处理，通过'"%d":%d'，'"%d":"%d"'格式化数组
# #                 ap.append('"%d":%d' % (k, v))
# #             else:
# #                 ap.append('"%d":"%d"' % (k, v)) 
# #         s = '{%d}' % (','.join(ap))   #继续格式化
# #         p.append(s)
# #     t ='[%d]' % (','.join(p)) #格式化
# #     print (t)
# #     with open('student4.json',"w") as f:
# #         f.write(t)
get_value()